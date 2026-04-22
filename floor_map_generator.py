"""
2Fフロアマップ生成スクリプト
- 配置図ファイル：セルオフセットで階段効果を表現
- データシートファイル：台番号一覧表
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os

class FloorMapGenerator:
    def __init__(self, output_dir="."):
        self.output_dir = output_dir
        self.gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')

    def create_floor_map_2f(self):
        """2Fの配置図を生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "2F配置図"

        # セル幅・高さ設定
        for col in range(1, 30):
            ws.column_dimensions[get_column_letter(col)].width = 8
        for row in range(1, 150):
            ws.row_dimensions[row].height = 20

        current_row = 1

        # ===== 右上：長方形ブロック =====
        # 2021-2031（1行・11台）
        current_row = self._add_block_2f(ws, current_row, 1, [2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031], "2021-2031")
        current_row += 2

        # 2032-2042 / 2077-2087（2行・各11台）
        current_row = self._add_block_2f(ws, current_row, 1, [2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042], "2032-2042")
        current_row = self._add_block_2f(ws, current_row, 1, [2077, 2078, 2079, 2080, 2081, 2082, 2083, 2084, 2085, 2086, 2087], "2077-2087")
        current_row += 2

        # 2088-2098 / 2121-2131（2行・各11台）
        current_row = self._add_block_2f(ws, current_row, 1, [2088, 2089, 2090, 2091, 2092, 2093, 2094, 2095, 2096, 2097, 2098], "2088-2098")
        current_row = self._add_block_2f(ws, current_row, 1, [2121, 2122, 2123, 2124, 2125, 2126, 2127, 2128, 2129, 2130, 2131], "2121-2131")
        current_row += 2

        # 2132-2142 / 2151-2161（2行・各11台）
        current_row = self._add_block_2f(ws, current_row, 1, [2132, 2133, 2134, 2135, 2136, 2137, 2138, 2139, 2140, 2141, 2142], "2132-2142")
        current_row = self._add_block_2f(ws, current_row, 1, [2151, 2152, 2153, 2154, 2155, 2156, 2157, 2158, 2159, 2160, 2161], "2151-2161")
        current_row += 3

        # ===== 右側：右斜め階段（7列） =====
        # 7列の台番号リスト
        columns_data = [
            [2162, 2163, 2164, 2165, 2166],  # 5台
            [2187, 2188, 2189, 2190],         # 4台
            [2191, 2192, 2193, 2194, 2195, 2196, 2197, 2198, 2199, 2200, 2201, 2202],  # 12台
            [2223, 2224, 2225, 2226, 2227, 2228, 2229, 2230, 2231, 2232, 2233, 2234, 2235],  # 13台
            [2236, 2237, 2238, 2239, 2240, 2241, 2242, 2243, 2244, 2245, 2246, 2247, 2248, 2249, 2250, 2251, 2252, 2253, 2254, 2255],  # 20台
            [2282, 2283, 2284, 2285, 2286, 2287, 2288, 2289, 2290, 2291, 2292, 2293, 2294, 2295, 2296, 2297, 2298, 2299, 2300, 2301],  # 20台
            [2302, 2303, 2304, 2305, 2306, 2307, 2308, 2309, 2310, 2311, 2312, 2313, 2314, 2315, 2316],  # 15台
        ]

        current_row = self._add_stair_columns(ws, current_row, columns_data)
        current_row += 3

        # ===== 下部：長方形ブロック =====
        # 2167-2176 / 2177-2186（2行・各10台）
        current_row = self._add_block_2f(ws, current_row, 1, [2167, 2168, 2169, 2170, 2171, 2172, 2173, 2174, 2175, 2176], "2167-2176")
        current_row = self._add_block_2f(ws, current_row, 1, [2177, 2178, 2179, 2180, 2181, 2182, 2183, 2184, 2185, 2186], "2177-2186")
        current_row += 2

        # 2203-2212 / 2213-2222（2行・各10台）
        current_row = self._add_block_2f(ws, current_row, 1, [2203, 2204, 2205, 2206, 2207, 2208, 2209, 2210, 2211, 2212], "2203-2212")
        current_row = self._add_block_2f(ws, current_row, 1, [2213, 2214, 2215, 2216, 2217, 2218, 2219, 2220, 2221, 2222], "2213-2222")
        current_row += 2

        # 2256-2268 / 2269-2281（2行・各13台）
        current_row = self._add_block_2f(ws, current_row, 1, [2256, 2257, 2258, 2259, 2260, 2261, 2262, 2263, 2264, 2265, 2266, 2267, 2268], "2256-2268")
        current_row = self._add_block_2f(ws, current_row, 1, [2269, 2270, 2271, 2272, 2273, 2274, 2275, 2276, 2277, 2278, 2279, 2280, 2281], "2269-2281")
        current_row += 2

        # 2317-2330（1行・14台）
        current_row = self._add_block_2f(ws, current_row, 1, [2317, 2318, 2319, 2320, 2321, 2322, 2323, 2324, 2325, 2326, 2327, 2328, 2329, 2330], "2317-2330")

        output_path = os.path.join(self.output_dir, "floor_map_2f.xlsx")
        wb.save(output_path)
        print(f"[OK] 配置図ファイル生成: {output_path}")
        return output_path

    def _add_block_2f(self, ws, start_row, start_col, machine_numbers, label=""):
        """台番号ブロック（横並び）を追加"""
        # 機種名スペース行（グレー塗り）
        for i, machine_num in enumerate(machine_numbers):
            cell = ws.cell(row=start_row, column=start_col + i)
            cell.fill = self.gray_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align

        # 台番号行
        for i, machine_num in enumerate(machine_numbers):
            cell = ws.cell(row=start_row + 1, column=start_col + i)
            cell.value = machine_num
            cell.border = self.thin_border
            cell.alignment = self.center_align
            cell.font = Font(size=9)

        return start_row + 2

    def _add_stair_columns(self, ws, start_row, columns_data):
        """右斜め階段（7列）を追加 - セルオフセット方式"""
        col_offset = 1
        current_col = col_offset

        for col_index, machines in enumerate(columns_data):
            # 各列の開始行をずらす（オフセット）
            row_offset = col_index * 3  # 3行ずつオフセット
            current_row = start_row + row_offset

            # 各台番号を縦に配置
            for machine_num in machines:
                # 機種名スペース（グレー塗り）
                cell = ws.cell(row=current_row, column=current_col)
                cell.fill = self.gray_fill
                cell.border = self.thin_border
                cell.alignment = self.center_align

                # 台番号
                cell = ws.cell(row=current_row + 1, column=current_col)
                cell.value = machine_num
                cell.border = self.thin_border
                cell.alignment = self.center_align
                cell.font = Font(size=8)

                current_row += 2

            current_col += 1

        # 最終行を返す
        return start_row + (len(columns_data) - 1) * 3 + max(len(col) for col in columns_data) * 2

    def create_data_sheet(self):
        """台番号一覧表を生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "台番号一覧"

        # ヘッダー
        headers = ["列名", "台数", "台番号一覧"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = self.center_align

        # カラム幅
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50

        # データ定義
        data = [
            ("2021-2031", 11, "2021～2031"),
            ("2032-2042", 11, "2032～2042"),
            ("2077-2087", 11, "2077～2087"),
            ("2088-2098", 11, "2088～2098"),
            ("2121-2131", 11, "2121～2131"),
            ("2132-2142", 11, "2132～2142"),
            ("2151-2161", 11, "2151～2161"),
            ("2162-2166", 5, "2162～2166"),
            ("2167-2176", 10, "2167～2176"),
            ("2177-2186", 10, "2177～2186"),
            ("2187-2190", 4, "2187～2190"),
            ("2191-2202", 12, "2191～2202"),
            ("2203-2212", 10, "2203～2212"),
            ("2213-2222", 10, "2213～2222"),
            ("2223-2235", 13, "2223～2235"),
            ("2236-2255", 20, "2236～2255"),
            ("2256-2268", 13, "2256～2268"),
            ("2269-2281", 13, "2269～2281"),
            ("2282-2301", 20, "2282～2301"),
            ("2302-2316", 15, "2302～2316"),
            ("2317-2330", 14, "2317～2330"),
        ]

        for row_idx, (col_name, count, machines) in enumerate(data, 2):
            ws.cell(row=row_idx, column=1).value = col_name
            ws.cell(row=row_idx, column=2).value = count
            ws.cell(row=row_idx, column=3).value = machines

        output_path = os.path.join(self.output_dir, "floor_data_2f.xlsx")
        wb.save(output_path)
        print(f"[OK] データシートファイル生成: {output_path}")
        return output_path

if __name__ == "__main__":
    generator = FloorMapGenerator()
    generator.create_floor_map_2f()
    generator.create_data_sheet()
    print("\n[OK] 生成完了")
